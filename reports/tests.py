from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .models import CarRecords, Driver


class ExcelExportTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.manager = User.objects.create_user(
            username="manager",
            password="password123",
            is_staff=True,
        )
        self.driver = Driver.objects.create(name="سائق الاختبار", id_number="123456789")
        CarRecords.objects.create(
            day="Monday",
            date="2026-01-01",
            name=self.driver,
            assistant="مساعد 1",
            car_count=3,
            cups=10,
            documented_count=2,
            documented_for="مؤسسة أ",
            notes="داخل الفلتر",
        )
        CarRecords.objects.create(
            day="Tuesday",
            date="2026-02-01",
            name=self.driver,
            assistant="مساعد 2",
            car_count=5,
            cups=20,
            documented_count=4,
            documented_for="مؤسسة ب",
            notes="خارج الفلتر",
        )

    def test_export_requires_login(self):
        response = self.client.get(reverse("export_car_records_excel"))

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])

    def test_car_records_export_is_xlsx_and_respects_date_filter(self):
        self.client.force_login(self.manager)

        response = self.client.get(
            reverse("export_car_records_excel"),
            {"start_date": "2026-01-01", "end_date": "2026-01-31"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["سجلات السيارات"]

        self.assertTrue(sheet.sheet_view.rightToLeft)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["A1"].value, "اليوم")
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][7], "مؤسسة أ")
        self.assertEqual(rows[0][8], "داخل الفلتر")

    def test_comprehensive_export_contains_expected_sheets(self):
        self.client.force_login(self.manager)

        response = self.client.get(reverse("download_comprehensive_excel"))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("الموظفون", workbook.sheetnames)
        self.assertIn("سجلات السيارات", workbook.sheetnames)
        self.assertIn("أماكن المصروفات", workbook.sheetnames)
