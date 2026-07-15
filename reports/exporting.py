from datetime import datetime, time, timedelta
from io import BytesIO

from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    Attendance,
    CarRecords,
    Device,
    Driver,
    Expenses,
    Expenses_type,
    FountainRecords,
    FuelTransaction,
    Institution,
    MainItem,
    Maintenance,
    MaintenanceLocation,
    Place_Expenses,
    SubItem,
    VehicleFuelRecord,
    Worker,
)


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ARABIC_DAYS = {
    "Saturday": "السبت",
    "Sunday": "الأحد",
    "Monday": "الاثنين",
    "Tuesday": "الثلاثاء",
    "Wednesday": "الأربعاء",
    "Thursday": "الخميس",
    "Friday": "الجمعة",
}


def empty(value, default="-"):
    return default if value in (None, "") else value


def format_date(value):
    if not value:
        return ""
    return value.strftime("%d-%m-%Y")


def format_time(value):
    if not value:
        return ""
    return value.strftime("%H:%M")


def display_choice(instance, field_name):
    getter = getattr(instance, f"get_{field_name}_display", None)
    if getter:
        return getter()
    return getattr(instance, field_name, "")


def dated_filename(prefix):
    today = timezone.localdate().strftime("%Y-%m-%d")
    return f"{prefix}_{today}.xlsx"


def workbook_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type=EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def create_workbook():
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def write_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title=title[:31])
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_index, row in enumerate(rows, 2):
        for column_index, value in enumerate(row, 1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = body_alignment

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 45)

    return sheet


def export_rows(filename_prefix, sheet_title, headers, rows):
    workbook = create_workbook()
    write_sheet(workbook, sheet_title, headers, rows)
    return workbook_response(workbook, dated_filename(filename_prefix))


def filter_workers(params):
    records = Worker.objects.all().order_by("name")
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    if start_date:
        records = records.filter(start_date__gte=start_date)
    if end_date:
        records = records.filter(start_date__lte=end_date)
    return records


def workers_rows(records):
    return [
        [
            worker.name,
            empty(worker.id_number),
            display_choice(worker, "job"),
            format_date(worker.start_date),
            worker.salary,
        ]
        for worker in records
    ]


def filter_attendance_report(params):
    employee_id = params.get("employee")
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    if not (employee_id and start_date and end_date):
        return Attendance.objects.none()
    return Attendance.objects.select_related("employee").filter(
        employee_id=employee_id,
        date__range=[start_date, end_date],
    ).order_by("date")


def attendance_report_rows(records):
    official_start_time = time(6, 0)
    official_end_time = time(15, 0)
    rows = []
    for record in records:
        delay = 0
        early_leave = 0
        if record.check_in and record.check_in > official_start_time:
            delay_timedelta = datetime.combine(datetime.today(), record.check_in) - datetime.combine(
                datetime.today(), official_start_time
            )
            delay = int(delay_timedelta.total_seconds() // 60)
        if record.check_out and record.check_out < official_end_time:
            early_leave_timedelta = datetime.combine(datetime.today(), official_end_time) - datetime.combine(
                datetime.today(), record.check_out
            )
            early_leave = int(early_leave_timedelta.total_seconds() // 60)
        rows.append(
            [
                record.employee.name,
                format_date(record.date),
                format_time(record.check_in),
                format_time(record.check_out),
                delay,
                early_leave,
            ]
        )
    return rows


def workers_delay_rows(params):
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    if not (start_date and end_date):
        return []

    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    total_days = (end - start).days + 1
    rows = []

    for employee in Worker.objects.all().order_by("name"):
        total_delay = 0
        total_early_leave = 0
        absence_days = 0
        attendances = Attendance.objects.filter(employee=employee, date__range=[start, end])
        attendance_by_date = {attendance.date: attendance for attendance in attendances}

        for index in range(total_days):
            current_day = start + timedelta(days=index)
            attendance = attendance_by_date.get(current_day)
            if attendance and attendance.check_in and attendance.check_out:
                expected_start = time(7, 0)
                expected_end = time(17, 0)
                if attendance.check_in > expected_start:
                    delay = datetime.combine(attendance.date, attendance.check_in) - datetime.combine(
                        attendance.date, expected_start
                    )
                    total_delay += delay.total_seconds() / 60
                if attendance.check_out < expected_end:
                    early = datetime.combine(attendance.date, expected_end) - datetime.combine(
                        attendance.date, attendance.check_out
                    )
                    total_early_leave += early.total_seconds() / 60
            else:
                absence_days += 1

        rows.append([employee.name, int(total_delay), int(total_early_leave), absence_days])
    return rows


def filter_fuel_transactions(params):
    records = FuelTransaction.objects.select_related("driver").all()
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    fuel_type = params.get("fuel_type")
    usage_type = params.get("usage_type")
    transaction_type = params.get("transaction_type")
    if start_date and end_date:
        records = records.filter(date__range=[start_date, end_date])
    if fuel_type:
        records = records.filter(fuel_type=fuel_type)
    if usage_type:
        records = records.filter(usage_type=usage_type)
    if transaction_type:
        records = records.filter(type=transaction_type)
    return records.order_by("date", "id")


def processed_fuel_rows(records):
    rows = []
    previous_meter_by_driver = {}
    previous_date_by_driver = {}

    for record in records:
        previous_meter = ""
        meter_difference = ""
        previous_date = ""
        date_difference = ""
        driver_id = record.driver_id

        if record.type == "out" and driver_id and record.meter_reading is not None:
            previous_meter = previous_meter_by_driver.get(driver_id, "")
            previous_date_value = previous_date_by_driver.get(driver_id)
            if previous_meter != "":
                meter_difference = record.meter_reading - previous_meter
            if previous_date_value is not None:
                previous_date = format_date(previous_date_value)
                date_difference = (record.date - previous_date_value).days
            previous_meter_by_driver[driver_id] = record.meter_reading
            previous_date_by_driver[driver_id] = record.date

        rows.append(
            [
                format_date(record.date),
                ARABIC_DAYS.get(record.date.strftime("%A"), record.date.strftime("%A")),
                display_choice(record, "type"),
                display_choice(record, "fuel_type"),
                record.quantity,
                display_choice(record, "fuel_transaction_source"),
                display_choice(record, "usage_type"),
                record.driver.name if record.driver else "",
                empty(record.vehicle_type, ""),
                empty(record.notes, ""),
                format_time(record.start_time),
                format_time(record.end_time),
                record.meter_reading if record.meter_reading is not None else "",
                previous_meter,
                meter_difference,
                previous_date,
                date_difference,
                record.total_cost if record.total_cost is not None else "",
            ]
        )

    rows.reverse()
    return rows


def filter_car_records(params):
    records = CarRecords.objects.select_related("name").all()
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    driver_id = params.get("driver")
    assistant_id = params.get("assistant")
    documented_for = params.get("documented_for")
    notes = params.get("notes")
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if driver_id:
        records = records.filter(name_id=driver_id)
    if assistant_id:
        assistant_name = Worker.objects.filter(id=assistant_id).values_list("name", flat=True).first()
        if assistant_name:
            records = records.filter(assistant__icontains=assistant_name)
    if documented_for:
        records = records.filter(documented_for__icontains=documented_for)
    if notes:
        records = records.filter(notes__icontains=notes)
    return records.order_by("-date")


def car_record_rows(records):
    return [
        [
            record.day,
            format_date(record.date),
            record.name.name if record.name else "",
            record.assistant,
            record.car_count,
            record.cups,
            record.documented_count,
            empty(record.documented_for, ""),
            empty(record.notes, ""),
        ]
        for record in records
    ]


def filter_fountain_records(params):
    records = FountainRecords.objects.all()
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    day = params.get("day")
    notes = params.get("notes")
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if day:
        records = records.filter(day__icontains=day)
    if notes:
        records = records.filter(notes__icontains=notes)
    return records.order_by("-date")


def fountain_record_rows(records):
    return [[record.day, format_date(record.date), record.car_count, empty(record.notes, "")] for record in records]


def filter_maintenance(params):
    records = Maintenance.objects.select_related("maintenance_location", "main_item", "sub_item", "device").all()
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    maintenance_location = params.get("maintenance_location")
    main_item = params.get("main_item")
    sub_item = params.get("sub_item")
    device = params.get("device")
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if maintenance_location:
        records = records.filter(maintenance_location_id=maintenance_location)
    if main_item:
        records = records.filter(main_item_id=main_item)
    if sub_item:
        records = records.filter(sub_item_id=sub_item)
    if device:
        records = records.filter(device_id=device)
    return records.order_by("-date", "-id")


def maintenance_rows(records):
    return [
        [
            record.day,
            format_date(record.date),
            str(record.maintenance_location or ""),
            str(record.main_item or ""),
            str(record.sub_item or ""),
            str(record.device or ""),
            empty(record.notes, ""),
            record.amount if record.amount is not None else "",
        ]
        for record in records
    ]


def filter_expenses(params):
    records = Expenses.objects.select_related("palce", "expenses_type").all()
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    place = params.get("place")
    expenses_type = params.get("expenses_type")
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if place:
        records = records.filter(palce_id=place)
    if expenses_type:
        records = records.filter(expenses_type_id=expenses_type)
    return records.order_by("-date", "-id")


def expenses_rows(records):
    return [
        [
            record.day,
            format_date(record.date),
            str(record.palce or ""),
            str(record.expenses_type or ""),
            record.amount if record.amount is not None else "",
            empty(record.notes, ""),
        ]
        for record in records
    ]


def filter_vehicle_fuel(params):
    records = VehicleFuelRecord.objects.select_related("driver").all()
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    driver = params.get("driver")
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if driver:
        records = records.filter(driver_id=driver)
    return records.order_by("-date", "-id")


def vehicle_fuel_rows(records):
    return [
        [
            format_date(record.date),
            record.driver.name if record.driver else "",
            record.fuel_quantity,
            record.odometer_before,
            record.odometer_after,
            record.distance_difference,
            empty(record.notes, ""),
        ]
        for record in records
    ]


def name_rows(records):
    return [[record.name] for record in records]


def financial_report_rows(params):
    from django.db.models.functions import TruncDay, TruncMonth, TruncYear

    start_date = params.get("start_date")
    end_date = params.get("end_date")
    report_type = params.get("report_type")
    trunc_fn = {"daily": TruncDay, "monthly": TruncMonth, "yearly": TruncYear}.get(report_type, TruncDay)

    fuel_qs = FuelTransaction.objects.all()
    worker_qs = Worker.objects.all()
    maintenance_qs = Maintenance.objects.all()
    expenses_qs = Expenses.objects.all()
    if start_date and end_date:
        fuel_qs = fuel_qs.filter(date__gte=start_date, date__lte=end_date)
        worker_qs = worker_qs.filter(start_date__gte=start_date, start_date__lte=end_date)
        maintenance_qs = maintenance_qs.filter(date__gte=start_date, date__lte=end_date)
        expenses_qs = expenses_qs.filter(date__gte=start_date, date__lte=end_date)

    fuel_data = fuel_qs.annotate(period=trunc_fn("date")).values("period").annotate(total_fuel=Sum("total_cost"))
    worker_data = worker_qs.annotate(period=trunc_fn("start_date")).values("period").annotate(total_worker=Sum("salary"))
    maintenance_data = maintenance_qs.annotate(period=trunc_fn("date")).values("period").annotate(total_maintenance=Sum("amount"))
    expenses_data = expenses_qs.annotate(period=trunc_fn("date")).values("period").annotate(total_expenses=Sum("amount"))

    all_periods = (
        set(fuel_data.values_list("period", flat=True))
        | set(worker_data.values_list("period", flat=True))
        | set(maintenance_data.values_list("period", flat=True))
        | set(expenses_data.values_list("period", flat=True))
    )
    rows = []
    for period in sorted([period for period in all_periods if period is not None]):
        fuel_total = next((item["total_fuel"] for item in fuel_data if item["period"] == period), 0) or 0
        worker_total = next((item["total_worker"] for item in worker_data if item["period"] == period), 0) or 0
        maintenance_total = next((item["total_maintenance"] for item in maintenance_data if item["period"] == period), 0) or 0
        expenses_total = next((item["total_expenses"] for item in expenses_data if item["period"] == period), 0) or 0
        rows.append(
            [
                format_date(period),
                fuel_total,
                worker_total,
                maintenance_total,
                expenses_total,
                fuel_total + worker_total + maintenance_total + expenses_total,
            ]
        )
    return rows


WORKER_HEADERS = ["الاسم", "رقم الهوية", "طبيعة العمل", "بداية العمل", "الراتب"]
ATTENDANCE_HEADERS = ["الموظف", "التاريخ", "وقت الحضور", "وقت الانصراف", "دقائق التأخير", "دقائق الانصراف المبكر"]
WORKERS_DELAY_HEADERS = ["الموظف", "إجمالي التأخير بالدقائق", "إجمالي الانصراف المبكر بالدقائق", "أيام الغياب"]
DRIVER_HEADERS = ["اسم السائق", "رقم الهوية"]
INSTITUTION_HEADERS = ["اسم المؤسسة"]
FUEL_HEADERS = [
    "التاريخ",
    "اليوم",
    "نوع العملية",
    "نوع المحروقات",
    "الكمية",
    "مصدر المحروقات",
    "الجهة",
    "السائق",
    "نوع السيارة",
    "ملاحظات",
    "وقت التشغيل",
    "وقت الإيقاف",
    "القراءة الحالية",
    "القراءة السابقة",
    "فرق العداد",
    "تاريخ القراءة السابقة",
    "فرق الأيام",
    "التكلفة",
]
CAR_HEADERS = ["اليوم", "التاريخ", "اسم السائق", "اسم المساعد", "عدد السيارات", "عدد الأكواب", "عدد الموثقة", "تم التوثيق لـ", "ملاحظات"]
FOUNTAIN_HEADERS = ["اليوم", "التاريخ", "عدد السيارات", "ملاحظات"]
MAINTENANCE_HEADERS = ["اليوم", "التاريخ", "مكان الصيانة", "البند الرئيسي", "البند الفرعي", "الجهاز", "الملاحظات", "المبلغ"]
EXPENSES_HEADERS = ["اليوم", "التاريخ", "مكان الصرف", "نوع الصرف", "المبلغ", "الملاحظات"]
VEHICLE_FUEL_HEADERS = ["التاريخ", "السائق", "كمية السولار", "قراءة قبل", "قراءة بعد", "الفرق", "ملاحظات"]
NAME_HEADERS = ["الاسم"]
FINANCIAL_HEADERS = ["الفترة", "إجمالي المحروقات", "إجمالي الرواتب", "إجمالي الصيانة", "إجمالي المصروفات", "الإجمالي"]


def single_report_response(report_key, params):
    reports = {
        "workers": ("workers_export", "الموظفون", WORKER_HEADERS, workers_rows(filter_workers(params))),
        "attendance": ("attendance_export", "الحضور والانصراف", ATTENDANCE_HEADERS, attendance_report_rows(filter_attendance_report(params))),
        "workers_delays": ("workers_delays_export", "تأخيرات العمال", WORKERS_DELAY_HEADERS, workers_delay_rows(params)),
        "drivers": ("drivers_export", "السائقون", DRIVER_HEADERS, [[driver.name, driver.id_number] for driver in Driver.objects.all().order_by("name")]),
        "institutions": ("institutions_export", "المؤسسات", INSTITUTION_HEADERS, name_rows(Institution.objects.all().order_by("name"))),
        "fuel": ("fuel_transactions_export", "حركة المحروقات", FUEL_HEADERS, processed_fuel_rows(filter_fuel_transactions(params))),
        "cars": ("car_records_export", "سجلات السيارات", CAR_HEADERS, car_record_rows(filter_car_records(params))),
        "fountain": ("fountain_records_export", "سجلات النافورة", FOUNTAIN_HEADERS, fountain_record_rows(filter_fountain_records(params))),
        "maintenance": ("maintenance_export", "الصيانة", MAINTENANCE_HEADERS, maintenance_rows(filter_maintenance(params))),
        "expenses": ("expenses_export", "المصروفات", EXPENSES_HEADERS, expenses_rows(filter_expenses(params))),
        "vehicle_fuel": ("vehicle_fuel_export", "وقود السيارات", VEHICLE_FUEL_HEADERS, vehicle_fuel_rows(filter_vehicle_fuel(params))),
        "financial": ("financial_report_export", "التقرير المالي", FINANCIAL_HEADERS, financial_report_rows(params)),
    }
    filename_prefix, sheet_title, headers, rows = reports[report_key]
    return export_rows(filename_prefix, sheet_title, headers, rows)


def comprehensive_workbook_response():
    workbook = create_workbook()
    sheets = [
        ("الموظفون", WORKER_HEADERS, workers_rows(Worker.objects.all().order_by("name"))),
        ("الحضور", ATTENDANCE_HEADERS, attendance_report_rows(Attendance.objects.select_related("employee").all().order_by("date", "id"))),
        ("السائقون", DRIVER_HEADERS, [[driver.name, driver.id_number] for driver in Driver.objects.all().order_by("name")]),
        ("المؤسسات", INSTITUTION_HEADERS, name_rows(Institution.objects.all().order_by("name"))),
        ("المحروقات", FUEL_HEADERS, processed_fuel_rows(FuelTransaction.objects.select_related("driver").all().order_by("date", "id"))),
        ("سجلات السيارات", CAR_HEADERS, car_record_rows(CarRecords.objects.select_related("name").all().order_by("-date"))),
        ("سجلات النافورة", FOUNTAIN_HEADERS, fountain_record_rows(FountainRecords.objects.all().order_by("-date"))),
        ("الصيانة", MAINTENANCE_HEADERS, maintenance_rows(Maintenance.objects.select_related("maintenance_location", "main_item", "sub_item", "device").all())),
        ("المصروفات", EXPENSES_HEADERS, expenses_rows(Expenses.objects.select_related("palce", "expenses_type").all())),
        ("وقود السيارات", VEHICLE_FUEL_HEADERS, vehicle_fuel_rows(VehicleFuelRecord.objects.select_related("driver").all().order_by("-date", "-id"))),
        ("أماكن الصيانة", NAME_HEADERS, name_rows(MaintenanceLocation.objects.all().order_by("name"))),
        ("الأصناف الرئيسية", NAME_HEADERS, name_rows(MainItem.objects.all().order_by("name"))),
        ("الأصناف الفرعية", NAME_HEADERS, name_rows(SubItem.objects.all().order_by("name"))),
        ("الأجهزة", NAME_HEADERS, name_rows(Device.objects.all().order_by("name"))),
        ("أنواع المصروفات", NAME_HEADERS, name_rows(Expenses_type.objects.all().order_by("name"))),
        ("أماكن المصروفات", NAME_HEADERS, name_rows(Place_Expenses.objects.all().order_by("name"))),
    ]
    for title, headers, rows in sheets:
        write_sheet(workbook, title, headers, rows)
    return workbook_response(workbook, dated_filename("comprehensive_export"))
